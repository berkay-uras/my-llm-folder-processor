from fastapi import FastAPI, UploadFile, File, HTTPException
from fastapi.responses import JSONResponse
from pydantic import BaseModel
from fastapi.middleware.cors import CORSMiddleware
import os
import io
import pypdf
from docx import Document
import openpyxl
import logging
import hashlib
from typing import List, Dict, Any, Optional
from contextlib import asynccontextmanager

# sqlite3 sürüm sorununu çözmek için eklenen kod
__import__('pysqlite3')
import sys
sys.modules['sqlite3'] = sys.modules.pop('pysqlite3')

# Proje kök dizinini Python yoluna ekleyerek modül bulma sorunlarını çözme
import os
import sys
# Eğer script backend/main.py içinde çalışıyorsa, bir üst dizini (projenin kök dizini) ekle
if 'backend' in os.path.basename(os.getcwd()):
    sys.path.append(os.path.abspath(os.path.join(os.getcwd(), '..')))
else:
    sys.path.append(os.path.abspath(os.getcwd()))

from fastapi import FastAPI, HTTPException, Request
from fastapi.middleware.cors import CORSMiddleware
import logging
# Vektör veritabanı ve embedding için yeni kütüphaneler
import chromadb
from chromadb.utils import embedding_functions
from sentence_transformers import SentenceTransformer

# MongoDB için
from pymongo import MongoClient
from datetime import datetime

# LLM Modeli
from llama_cpp import Llama

# Logging configuration
logger = logging.getLogger(__name__)
logger.setLevel(logging.INFO)
handler = logging.StreamHandler()
formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
handler.setFormatter(formatter)
if not logger.handlers:
    logger.addHandler(handler)


# --- CONFIGURATION (Ayarlarınızı burada tanımlayın) ---
class Settings:
    LLM_MODELS_PATH = os.path.join("backend", "models")
    
    # Yeni: Model dosya adını ortam değişkeninden oku, yoksa varsayılanı kullan
    LLM_MODEL_FILENAME = os.getenv(
        "LLAMA_MODEL_FILENAME", 
        "mistral-7b-instruct-v0.2.Q4_K_M.gguf"
    )

    SUPPORTED_MODELS = {
        "mistral-7b": {
            # Yeni: Dosya adını ortam değişkeninden alınan değer ile kullan
            "file": LLM_MODEL_FILENAME,
            "llama_params": {
                "n_ctx": 2048,  # Modelin bağlam boyutunu ayarlayın
                "n_gpu_layers": -1,
                "use_mlock": False,
            }
        },
    }

settings = Settings()


# --- LLM Service (Sizin paylaştığınız kod) ---
class LLMService:
    def __init__(self):
        self.models: Dict[str, Llama] = {}
        self._load_models()

    def _load_models(self):
        logger.info(f"Loading LLM models from: {settings.LLM_MODELS_PATH}")
        # Yeni: Sadece tek bir modeli yüklüyoruz ve dosya adını settings'ten alıyoruz
        model_key = "mistral-7b"
        model_config = settings.SUPPORTED_MODELS[model_key]
        model_file = model_config["file"]
        llama_params = model_config.get("llama_params", {})
        
        model_path = os.path.join(settings.LLM_MODELS_PATH, model_file)
        if os.path.exists(model_path):
            logger.info(f"Attempting to load model: {model_path} with params: {llama_params}")
            try:
                self.models[model_key] = Llama(
                    model_path=model_path,
                    **llama_params
                )
                logger.info(f"Successfully loaded model: {model_key}")
            except Exception as e:
                logger.error(f"Error loading model {model_key} from {model_path}: {e}")
        else:
            logger.warning(f"Model file not found: {model_path}")

    def get_supported_models(self) -> List[str]:
        return list(self.models.keys())

    def get_llm_response(self, model_name: str, prompt: str, document_content: Optional[str] = None) -> str:
        if model_name not in self.models:
            raise ValueError(f"Model '{model_name}' not supported or not loaded.")

        llm = self.models[model_name]
        
        # Bu LLM Service'i bizim projemize uyarlamak için prompt yapısını değiştiriyoruz.
        # RAG için gerekli olan document_content'i kullanıyoruz.
        full_prompt = f"Using the following documents, answer the question. If the documents do not contain the answer, say 'I don't know'.\n\nDocuments:\n{document_content}\n\nQuestion: {prompt}\n\nAnswer:"
        
        # Default parameters for generation
        generate_params = {
            "max_tokens": 1000,
            "stop": ["Question:", "\n\n"],
            "echo": False
        }

        logger.info(f"Generating response for model '{model_name}' with prompt: '{prompt[:50]}...' and params: {generate_params}")

        try:
            output = llm(
                full_prompt,
                **generate_params
            )
            return output["choices"][0]["text"].strip()
        except Exception as e:
            logger.error(f"Error during LLM response generation for model '{model_name}': {e}")
            raise ValueError(f"LLM response generation failed: {e}")

llm_service = None


# --- GLOBAL VARIABLES & COMPONENT INITIALIZATION ---
# MongoDB Connection
MONGO_URI = os.getenv("MONGO_URI", "mongodb://localhost:27017/")
mongo_client = None
files_metadata_collection = None
try:
    mongo_client = MongoClient(MONGO_URI)
    db = mongo_client["llm_project_db"]
    files_metadata_collection = db["files_metadata"]
    logger.info("✅ MongoDB connection successful.")
except Exception as e:
    logger.error(f"❌ MongoDB connection error: {e}")

# ChromaDB Path & Client
CHROMA_DB_PATH = os.getenv("CHROMA_DB_PATH", "./chroma_data")
chroma_client = None
collection = None
embedding_function = None


@asynccontextmanager
async def lifespan(app: FastAPI):
    global llm_service, chroma_client, collection, embedding_function

    # --- Uygulama Başlangıç Kodları ---
    # LLMService'i burada başlatıyoruz, böylece model yükleme işlemi ilk açılışta gerçekleşir.
    llm_service = LLMService()
    
    # ChromaDB ve Embedding Modelini Başlat
    try:
        chroma_client = chromadb.PersistentClient(path=CHROMA_DB_PATH)
        embedding_function = embedding_functions.SentenceTransformerEmbeddingFunction(model_name="sentence-transformers/all-MiniLM-L6-v2")
        collection = chroma_client.get_or_create_collection(
            name="llm_project_collection",
            embedding_function=embedding_function
        )
        logger.info("✅ ChromaDB initialized.")
    except Exception as e:
        logger.error(f"❌ Error initializing ChromaDB: {e}")
        collection = None
    
    yield

    # --- Uygulama Kapanış Kodları ---
    logger.info("Application is shutting down. Performing cleanup...")
    if mongo_client:
        mongo_client.close()


app = FastAPI(lifespan=lifespan)

origins = ["http://localhost:3000", "http://127.0.0.1:3000"]
app.add_middleware(
    CORSMiddleware,
    allow_origins=origins,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


# --- HELPER FUNCTIONS ---
def get_file_hash(file_stream) -> str:
    hasher = hashlib.sha256()
    buf = file_stream.read(io.DEFAULT_BUFFER_SIZE)
    while len(buf) > 0:
        hasher.update(buf)
        buf = file_stream.read(io.DEFAULT_BUFFER_SIZE)
    return hasher.hexdigest()

def chunk_text(text: str, chunk_size: int = 512, chunk_overlap: int = 50) -> List[str]:
    chunks = []
    text_len = len(text)
    start = 0
    while start < text_len:
        end = start + chunk_size
        chunks.append(text[start:end])
        start += chunk_size - chunk_overlap
    return chunks

def read_txt_md_file(file_stream):
    try:
        file_stream.seek(0)
        return file_stream.read().decode('utf-8')
    except UnicodeDecodeError:
        try:
            file_stream.seek(0)
            return file_stream.read().decode('latin-1')
        except Exception:
            return None
    except Exception:
        return None

def read_pdf_file(file_stream):
    text = ""
    try:
        file_stream.seek(0)
        reader = pypdf.PdfReader(io.BytesIO(file_stream.read()))
        for page in reader.pages:
            text += page.extract_text() + "\n"
        return text
    except Exception:
        logger.error("Error reading PDF file", exc_info=True)
        return None

def read_docx_file(file_stream):
    text = ""
    try:
        file_stream.seek(0)
        doc = Document(io.BytesIO(file_stream.read()))
        for para in doc.paragraphs:
            text += para.text + "\n"
        return text
    except Exception:
        return None

def read_xlsx_file(file_stream):
    text = ""
    try:
        file_stream.seek(0)
        workbook = openpyxl.load_workbook(io.BytesIO(file_stream.read()))
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            text += f"--- Sheet: {sheet_name} ---\n"
            for row in sheet.iter_rows():
                row_values = []
                for cell in row:
                    if cell.value is not None:
                        row_values.append(str(cell.value))
                if row_values:
                    text += "\t".join(row_values) + "\n"
            text += "\n"
        return text
    except Exception:
        return None


# --- API ENDPOINTS ---
@app.post("/sync-and-process/")
async def sync_and_process_files(files: list[UploadFile] = File(...)):
    if not collection:
        raise HTTPException(status_code=500, detail="ChromaDB is not initialized.")
    if files_metadata_collection is None:
        raise HTTPException(status_code=500, detail="MongoDB is not initialized for metadata.")
    
    if not files:
        raise HTTPException(status_code=400, detail="No files uploaded.")

    current_filenames = {file.filename for file in files}
    indexed_files_cursor = files_metadata_collection.find({})
    indexed_files = {doc["filename"]: doc["file_hash"] for doc in indexed_files_cursor}

    files_to_delete = []
    files_to_update = []
    
    for filename in indexed_files.keys():
        if filename not in current_filenames:
            files_to_delete.append(filename)
    
    if files_to_delete:
        try:
            collection.delete(where={"source": {"$in": files_to_delete}})
            files_metadata_collection.delete_many({"filename": {"$in": files_to_delete}})
            logger.info(f"Successfully deleted data for files: {files_to_delete}")
        except Exception as e:
            logger.error(f"Error deleting data for files {files_to_delete}: {e}")

    processed_count = 0
    for file in files:
        file.file.seek(0)
        file_hash = get_file_hash(file.file)
        
        if indexed_files.get(file.filename) == file_hash:
            logger.info(f"File '{file.filename}' unchanged. Skipping.")
            continue
            
        collection.delete(where={"source": file.filename})
        files_metadata_collection.delete_one({"filename": file.filename})
        
        file_extension = os.path.splitext(file.filename)[1].lower()
        file_content = None
        if file_extension in ['.txt', '.md']:
            file_content = read_txt_md_file(file.file)
        elif file_extension == '.pdf':
            file_content = read_pdf_file(file.file)
        elif file_extension == '.docx':
            file_content = read_docx_file(file.file)
        elif file_extension == '.xlsx':
            file_content = read_xlsx_file(file.file)
        else:
            logger.warning(f"Unsupported file format skipped: {file.filename}")
            continue

        if not file_content:
            logger.error(f"Could not extract content from file: {file.filename}")
            continue

        chunks = chunk_text(file_content)
        chunk_ids = [f"{file.filename}_{i}" for i in range(len(chunks))]
        metadatas = [{"source": file.filename, "chunk_id": i} for i in range(len(chunks))]
        
        try:
            collection.add(documents=chunks, metadatas=metadatas, ids=chunk_ids)
            files_metadata_collection.insert_one({
                "filename": file.filename,
                "file_hash": file_hash,
                "uploaded_at": datetime.utcnow().isoformat(),
            })
            processed_count += 1
            logger.info(f"File '{file.filename}' indexed successfully. {len(chunks)} chunks added.")
        except Exception as e:
            logger.error(f"Error indexing file {file.filename}: {e}")
    
    deleted_count = len(files_to_delete)
    message = f"Process completed. {processed_count} new or updated files were indexed. {deleted_count} files were deleted."

    return JSONResponse(content={
        "message": message,
        "indexed_documents_count": collection.count(),
        "llm_response": "The knowledge base has been updated. You can now ask a question."
    })

# RAG ile LLM'e soru sorma endpoint'i
class QuestionRequest(BaseModel):
    question: str
    model_name: str = "mistral-7b" # Varsayılan olarak bu modeli kullanıyoruz

@app.post("/ask-llm/")
async def ask_llm(request: QuestionRequest):
    if not collection:
        raise HTTPException(status_code=500, detail="ChromaDB is not initialized.")
    
    # Yeni: LLM servisinin yüklenip yüklenmediğini kontrol et
    if not llm_service or request.model_name not in llm_service.models:
        raise HTTPException(status_code=503, detail="LLM service is not ready. Please wait a moment.")
    
    user_question = request.question
    
    results = collection.query(
        query_texts=[user_question],
        n_results=5
    )
    
    if not results or not results['documents'][0]:
        return JSONResponse(content={"question": user_question, "llm_answer": "I couldn't find any relevant information in the documents to answer your question."})

    retrieved_documents = " ".join(results['documents'][0])
    
    llm_response_text = "An error occurred while getting LLM response."
    try:
        # LLMService sınıfını kullanarak yanıtı alıyoruz
        llm_response_text = llm_service.get_llm_response(
            model_name=request.model_name,
            prompt=user_question,
            document_content=retrieved_documents
        )
        logger.info("LLM response successfully received.")
    except Exception as e:
        logger.error(f"Failed to get LLM response: {e}")
        llm_response_text = f"An error occurred: {e}"

    return JSONResponse(content={
        "question": user_question,
        "llm_answer": llm_response_text
    })

@app.get("/")
async def root():
    return {"message": "FastAPI LLM Project Backend is running with ChromaDB!"}
